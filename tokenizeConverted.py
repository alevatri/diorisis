import re
import os
import sys
import lxml.etree as document
from utf2beta import convertUTF
from beta2utf import convertBeta
import time
from openpyxl import load_workbook
import configparser

os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')
file_list = config['paths']['file_list']
corpus_location = config['paths']['raw_corpora']
tf = config['paths']['tokenized']
logf = config['paths']['logs']

os.system("clear && printf '\e[3J'")
attrName=['lemma', 'translation', 'morph']
errorLog = open('%s/tokenize_errors_converted.txt'%logf,'w')

####FUNCTIONS####
def tagWONS(item):
	tag = item.tag
	if tag is document.Comment: tag = 'comment'
	return re.sub('\{.*?\}', '', tag)

def tokenizePerseus(fileName, author, title):
	progPercent = 0
	fileName = '%s/%s'%(corpus_location, fileName)

	#load file
	try:
		parse = document.parse(fileName)
	except:
		removeEntities = open(fileName, 'r').read()
		removeEntities = removeEntities.replace("&ast;", "*")
		removeEntities = removeEntities.replace("&mdash;", " — ")
		removeEntities = re.sub('&.*?;', '', removeEntities)
		try:
			parse = document.fromstring(removeEntities)
		except:
			errorLog.write('Error parsing %s:\n|_____%s\n'%(fileName,sys.exc_info()[1]))
			return
		else:
			xmlroot = parse.getroottree()
	else:	
		xmlroot = parse.getroot()
		
	#extract metadata
	fileStr=open(fileName, 'r').read()
	if re.search('[α-ω]', fileStr) != None:
		isunicode = True
	else:
		isunicode = False
	del fileStr
	tlgId = xmlroot.xpath("//tlgId")[0].text
	tlgAuthor = xmlroot.xpath("//tlgAuthor")[0].text
	if os.path.exists('%s/%s - %s (%s).xml'%(tf, author, title, tlgId)):
		return
	print('\t%s (%s): %s (%s)'%(author,tlgAuthor,title,tlgId))
	
	#generate header

	outputRoot = document.Element("TEI.2")
	teiHeader = document.SubElement(outputRoot, "teiHeader")
	fileDesc = document.SubElement(teiHeader, "fileDesc")
	titleStmt = document.SubElement(fileDesc, "titleStmt")
	titleNode = document.SubElement(titleStmt, "title")
	authorNode = document.SubElement(titleStmt, "author")
	tlgAuthorNode = document.SubElement(titleStmt, "tlgAuthor")
	tlgIdNode = document.SubElement(titleStmt, "tlgId")

	authorNode.text = author
	titleNode.text = title
	tlgAuthorNode.text = tlgAuthor
	tlgIdNode.text = tlgId


	#annotation
	try:
		sourceText = xmlroot.xpath('//namespace:body', namespaces={'namespace' : 'http://www.tei-c.org/ns/1.0'})[0]
	except:
		sourceText = xmlroot.xpath('//body')[0]

	annoText = document.SubElement(outputRoot, "text")
	body = document.SubElement(annoText, "body")

	sentenceId = 0
	newSentence = True
	wordId = 0 # reset after each punctuation mark . · ;
	location = ""

	totalLength=len(sourceText.xpath(".//*"))
	index = 0
	brokenWord = ''
	for elem in sourceText.iter():
		index += 1
		if tagWONS(elem)=="div" or tagWONS(elem)=="l" or tagWONS(elem)=="quote" or tagWONS(elem)=="cit" or tagWONS(elem)=="add":
	
			#retrieve text and annotate quotes and lacunas
			nodesWithText='./text()|.//l/text()|.//quote/text()|.//add/text()|.//cit//text()|.//said/text()|.//q/text()|.//p/text()'
			nodesWithTextN='./text()|.//n:l/text()|.//n:quote/text()|.//n:add/text()|.//n:cit//text()|.//n:said/text()|.//n:q/text()|.//n:p/text()'
			rightNodes = elem.xpath(nodesWithText)
			rightNodesN = elem.xpath(nodesWithTextN, namespaces={'n' : 'http://www.tei-c.org/ns/1.0'})
			if len(rightNodesN) > len(rightNodes): rightNodes = rightNodesN
			for idx, snippet in enumerate(rightNodes):
				snippetAdd = False
				snippetQuote = False
				if (tagWONS(snippet.getparent()) == 'add' and snippet.is_text) or len(snippet.getparent().xpath('ancestor::add'))>0 or len(snippet.getparent().xpath('ancestor::n:add', namespaces={'n' : 'http://www.tei-c.org/ns/1.0'}))>0:
					snippetAdd = True
				if (tagWONS(snippet.getparent()) == 'quote' and snippet.is_text) or len(snippet.getparent().xpath('ancestor::quote'))>0 or len(snippet.getparent().xpath('ancestor::n:quote', namespaces={'n' : 'http://www.tei-c.org/ns/1.0'}))>0:
					snippetQuote = True
				if snippetAdd == True:
					snippet = "2%s"%snippet
					snippet = " 2".join(snippet.split())
				if snippetQuote == True:
					snippet = "3%s"%snippet
					snippet = " 3".join(snippet.split())
				rightNodes[idx] = snippet
			currentText = ' '.join(rightNodes)
			currentText = '%s%s'%(brokenWord,currentText)
			brokenWord = ''
			if re.search('[^\s]-$', currentText) != None:
				try:
					brokenWord = re.search('\s([^\s]+)-$', currentText).group(1)
					currentText = re.sub('\s[^\s]+-$', '', currentText)
				except:
					brokenWord = re.search('^([^\s]+)-$', currentText).group(1)
					currentText = re.sub('^[^\s]+-$', '', currentText)
			currentText = currentText.replace('--', ' — ')
			currentText = currentText.replace('- ', '-')
			currentText = currentText.replace('-', '')
				
			#retrieve location
			section = []
			iterator = 1
			while True:
				if eval('tagWONS(elem.%s)'%('.'.join(iterator*['getparent()']))) == 'body':
					break
				elif eval('elem.%sget("n")'%(iterator*'getparent().')) != None and eval('elem.%sget("n")'%(iterator*'getparent().'))[:3] != 'urn':
					section.append(eval('elem.%sget("n")'%(iterator*'getparent().')))
				iterator += 1
			section.reverse()
			location = ".".join(section)
			#if line number, use this as location
			if tagWONS(elem) == "l" and elem.get("n") != None:
				location = ".".join([location, elem.get("n")])
				if location[0] == ".": location = location[1:]
		
			#fix spacing of opening brackets
			currentText = currentText.replace("_", "_ ")
			if isunicode == True:
				currentText = currentText.replace("(", "( ")
			#fix curly quotes
			currentText = currentText.replace("”", '"').replace("“", '"')
			currentText = re.sub('([^\s])"([^\s])', r'\1 \2', currentText)
			currentText = currentText.replace("ʽ", "'").replace("̓", "'").replace("ʼ", "'")
		
			#parse words
			for words in currentText.split():
				if newSentence == True:
					sentenceId += 1
					document.SubElement(body, "sentence", id = str(sentenceId), location = str(location))
					newSentence = False	
				currentSentence = body.xpath('./sentence[@id="'+str(sentenceId)+'"]')[0]
				if words.startswith("(") and isunicode == True:
					document.SubElement(currentSentence, "punct", mark="(")
				elif words.startswith("—"):
					document.SubElement(currentSentence, "punct", mark="—")		
				else:								
					#search for punctuation marks
					word = words.replace("·", ":")
					if isunicode == True:
						word = re.sub('([,\.;:\_\)])', r' \1', word)
					else:
						word = re.sub('([,\.;:\_])', r' \1', word)
					punctuation = ""
					if len(word.split()) > 1:
						punctuation = word.split()[1].strip()
					
					#convert word to betacode
					if isunicode == True:
						word = convertUTF(word)
								
					#remove non-alphabetic characters
					word = re.sub('[^a-z\(\)\\\/\*\|=\+\'23]*', '', word)
					word = re.sub('(?<=[a-z\(\)\\\/\*\|=\+\'23])[^a-z\(\)\\\/\*\|=\+\'23]*(?=[a-z\(\)\\\/\*\|=\+\'23])', '', word)
					word = re.search('[a-z\(\)\\\/\*\|=\+\'23]*', word)
			
					#turn graves into acutes, lowercase betacode, remove clitic accent, escape diacritics
					wordForm = word.group(0).replace("2", "").replace("3", "")
					word = word.group(0).replace('\\', '/').lower()
					word = word.replace("/", "7").replace("\\", "8").replace("'", "£").replace("*", "~").replace("+", "€").replace("|", "9").replace("!", "ª")
					howmanyAccents = len(re.findall("[78=]", word))
					if howmanyAccents > 1:
						word=word[::-1].replace("7", "", 1)[::-1]

					#display progress
					progPercent = str(int(index*100/totalLength))
					sys.stdout.write("\r\033[K\tProgress: %s%%"%(progPercent))
					sys.stdout.flush()
																		
					#create word node
					if wordForm != '':
						currWord = document.SubElement(currentSentence, "word", form=wordForm)
						#add attribute if lacuna or quote and clear for lookup
						if '2' in word:
							currWord.set('lacuna', 'true')
							word=word.replace('2', '')
						if '3' in word:
							currWord.set('isquote', 'true')
							word=word.replace('3', '')
						#set word id
						wordId += 1
						currWord.set('id', str(wordId))	
	
					#append punctuation node, if mark present
					if len(punctuation) > 0 :
						newPunct = document.SubElement(currentSentence, "punct", mark=punctuation)
						if re.search(r'[\.;:]', punctuation) != None: 
							newSentence = True
							wordId = 0
					time.sleep(0.0001)	
			else:
				pass
	#output
	print()
	outputFile = document.ElementTree(outputRoot)
	outputFile.write('%s/%s - %s (%s).xml'%(tf, author, title, tlgId), xml_declaration = True, encoding='UTF-8', pretty_print=True)
	file[h['Tokenized file']].value = '%s - %s (%s).xml'%(author, title, tlgId)
	file[h['TLG Author']].value = tlgAuthor
	file[h['TLG ID']].value = tlgId
	
###################
wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws[config['excel_range']['range']]
index = 0
for file in files:
	if file[h['Source']].value == 'Perseus':
		continue
	index+=1
	print('%d out of %d'%(index,len(files))) 
	tokenizePerseus(file[h['Source file']].value, file[h['Author']].value, file[h['Work']].value)
	print()
errorLog.close()
wb.save('%s/file_list.xlsx'%file_list)