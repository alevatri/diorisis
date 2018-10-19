import re
import os
import sys
import lxml.etree as document
from utf2beta import convertUTF
from beta2utf import convertBeta
from grkFrm import greekForms 
from grkLemmata import greekLemmata
import time
from openpyxl import load_workbook
import configparser
from copy import deepcopy

os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')
file_list = config['paths']['file_list']
tf = config['paths']['tokenized']
af = config['paths']['annotated']
rc = config['paths']['raw_corpora']
logf = config['paths']['logs']
parser = document.XMLParser(remove_blank_text=True)

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws[config['excel_range']['range']]

os.system("clear && printf '\e[3J'")
attrName=['lemma', 'translation', 'morph']
unknownlist = open('%s/unknown_list.txt'%logf, 'a')
unknownlist_proper = open('%s/unknown_list_proper.txt'%logf, 'a')
unknownlist_single = open('%s/unknown_list_single.txt'%logf, 'a')
unknown_dict = {}

#####FUNCTIONS#####
def cleanWords(word):
	word = re.sub('[^a-z\(\)\\\/\*\|=\+\'0-9\"]*', '', word)

	#turn graves into acutes, lowercase betacode, remove clitic accent, escape diacritics
	wordForm = word.replace("2", "").replace("3", "")
	word = word.replace('\\', '/').lower()
	word = re.sub(r"\*([aehiouw])([\(\)/=]+)",r"*\2\1", word) #parole registrate maiuscole: *DIACRlettere; parole minuscole: *parola normale, quindi * + e)/ etc.
	word = re.sub(r"([\(\)/=]+)(\*)([aehiouw])",r"\2\1\3", word)
	word = re.sub(r"\|([/\(\)])",r"\g<1>|", word)
	howmanyAccents = len(re.findall("[/=]", word))
	if howmanyAccents > 1:
		word=word[::-1].replace("/", "", 1)[::-1]
	return word
	
def lookup(word):
	word = cleanWords(word)
	try:
		entry=greekForms[word]
	#no match			
	except KeyError:
		word = word.replace("+", "")
		word = re.sub(r"\*([\(\)/=]+)([aehiouw])",r"*\2\1", word)
		word2 = word.replace("'", "")
		notFound = False
		try:
			entry=greekForms[word]
		except KeyError:
			notFound = True
		else:
			return {'is_unknown':False, 'entry':entry}
		if notFound == True:
			word = word.replace("*", "")		
			try:
				entry=greekForms[word]
			except KeyError:
				word = re.sub('^c(?=u[nmg])','s', word)
				try:
					entry=greekForms[word]
				except KeyError:
					word = word.replace("'", "")
					try:
						entry=greekForms[word]
					except KeyError:
						word = re.sub("\)$", "'",word)
						try:
							entry=greekForms[word]
						except KeyError:
							word = re.sub("'$", "",word)
							try:
								entry=greekForms[word]
							except KeyError:
								word = re.sub("^\(", "",word)
								try:
									entry=greekForms[word]
								except KeyError:
									word = re.sub("\)$", "",word)
									try:
										entry=greekForms[word]
									except KeyError:
										word = re.sub("'", "",word)
										try:
											entry=greekForms[word]
										except KeyError:
											return {'is_unknown':True}
										else:
											return {'is_unknown':False, 'entry':entry}
									else:
										return {'is_unknown':False, 'entry':entry}
								else:
									return {'is_unknown':False, 'entry':entry}
							else:
								return {'is_unknown':False, 'entry':entry}
						else:
							return {'is_unknown':False, 'entry':entry}
					else:
						return {'is_unknown':False, 'entry':entry}
				else:
					return {'is_unknown':False, 'entry':entry}
			else:
					return {'is_unknown':False, 'entry':entry}
	else:
		return {'is_unknown':False, 'entry':entry}

def process(lemmata, word):
	#each wordform corresponds to a list of lemmata
	global wml
	global wml_mPos
	if len(lemmata)>1:
		wml+=1
		pos_in_word={}
		for lemma in lemmata:
			curr_pos = greekLemmata[lemma['l']]['pos']
			#score disambiguability (measure how many POS can a word be, and how many types per POS)
			pos_in_word[curr_pos]=pos_in_word.setdefault(curr_pos, 0)+1
		#disambiguability coefficient is given by the total number of POS represented only by one lemma divided by total lemmata (i.e. len(lemmata)). Assumption of equal probability of all lemmata.
		pos_one_lemma = 0
		for pos,count in pos_in_word.items():
			if count == 1: pos_one_lemma+=1
		dis_coefficient = pos_one_lemma/len(lemmata)
		#since the coefficient should be applied to each ambiguous word, it is equal to the disambiguability score (i.e. dis_score = dis_coefficient * 1)
		#the score should be subtracted from wml
		wml_mPos+=dis_coefficient
			
	for lemma in lemmata:
	#each lemma is a dictionary; 'l' = entry, 'a' is a list
		newLemma = document.SubElement(word, "lemma")
		newLemma.set('id',lemma['l'])
		newLemma.set('entry',greekLemmata[lemma['l']]['lemma'])
		newLemma.set('POS',greekLemmata[lemma['l']]['pos'])
		for analysis in lemma['a']:
			newAnalysis = document.SubElement(newLemma, "analysis")
			newAnalysis.set('morph', analysis)
			del newAnalysis
		del newLemma
		
###################

for record in files:
	file = '%s/%s'%(tf,record[h['Tokenized file']].value)
	unknowncount = 0
	unknowncount_proper = 0
	unknowncount_single = 0
	wml = 0 #words with multiple lemmata
	wml_mPos = 0 #word with multiple lemmata disambiguable by part of speech (cases: 1*POS1,1*POS2,1*...,1*POSn = T; >1*POS1,...,1*POSn = T only if tagger outputs POS occurring only once; >1*POS1,...,>1*POSn = F
	print('%s - %s'%(record[h['Author']].value,record[h['Work']].value))
	progPercent=0
	fileName = record[h['Tokenized file']].value
	if os.path.exists('%s/%s'%(af, fileName)):
		continue
	parse = document.parse(file, parser)

	#metadata
	parse.xpath('//tlgAuthor')[0].text = record[h['TLG Author']].value
	parse.xpath('//tlgId')[0].text = record[h['TLG ID']].value
	teiHeader = parse.xpath('//teiHeader')[0]
	sourceDesc = document.SubElement(teiHeader.xpath('./fileDesc')[0], 'sourceDesc')
	ref = document.SubElement(sourceDesc, 'ref', target = record[h['Source URL']].value)
	ref.text = record[h['Source']].value
	if record[h['Source']].value == 'Perseus':
		biblFull =  document.SubElement(sourceDesc, 'biblFull')
		path = '%s/%s'%(rc,record[h['Source file']].value)
		try:
			PerseusSource = document.parse(path, parser)
		except:
			removeEntities = open(path, 'r').read()
			removeEntities = removeEntities.replace("&ast;", "*")
			removeEntities = removeEntities.replace("&mdash;", " — ")
			removeEntities = re.sub('&.*?;', '', removeEntities)
			PerseusSource = document.fromstring(removeEntities, parser)
		PerseusHeader = PerseusSource.xpath('//fileDesc/*')
		if len(PerseusHeader) == 0:	PerseusHeader = PerseusSource.xpath('//*[local-name() = "fileDesc"]/*')
		[biblFull.append(deepcopy(x)) for x in PerseusHeader]
	profileDesc = document.SubElement(teiHeader, 'profileDesc')
	langUsage = document.SubElement(profileDesc, 'langUsage')
	language = document.SubElement(langUsage, 'language', ident='grc')
	language.text = 'Greek'
	creation = document.SubElement(profileDesc, 'creation')
	dateCr = document.SubElement(creation, 'date')
	dateCr.text = str(record[h['Date']].value)
	xenoData = document.SubElement(teiHeader, 'xenoData')
	genre = document.SubElement(xenoData, 'genre')
	genre.text = record[h['Genre']].value
	subgenre = document.SubElement(xenoData, 'subgenre')
	subgenre.text = record[h['Subgenre']].value	
	
	#parsing
	words = parse.xpath('//word')
	total = len(words)
	for idx, word in enumerate(words):
		progPercent = str(int(idx*100/total))
		sys.stdout.write("\r\033[K\tProgress: %s%%"%(progPercent))
		sys.stdout.flush()
		#get word form
		form = word.get("form")
		isunknown = False
		#lookup in dictionary
		if lookup(form)['is_unknown'] == True:
		#####test lacunae (merge with following words)#####
			try:
				next_word = words[idx+1]
			except:
				isunknown = True
			else:
				next_form = next_word.get("form")
				iter = 0
				while True:
					iter += 1
					previous_word = words[idx-iter]
					if previous_word.get('toremove') == None:
						break
				previous_form = previous_word.get("form")
				merged_form = ''
				if word.get('lacuna') != None:
					merged_form = '%s%s'%(form, next_form)
					if lookup(merged_form)['is_unknown'] == True:
						try:
							next_next_word = words[idx+2]
						except:
							isunknown = True
						else:
							next_next_form = next_next_word.get("form")
							if next_next_word.get('lacuna') != None:
								merged_form='%s%s'%(merged_form, next_next_form)
								if lookup(merged_form)['is_unknown'] == True:
									try:
										next_next_next_word = words[idx+3]
									except:
										isunknown = True
									else:
										next_next_next_form = next_next_next_word.get("form")
										merged_form = '%s%s'%(merged_form, next_next_next_form)
										if lookup(merged_form)['is_unknown'] == True:
											isunknown = True
										else:
											word.set('form', merged_form)
											process(lookup(merged_form)['entry'], word)
											next_word.set('toremove', 'true')
											next_next_word.set('toremove', 'true')
											next_next_next_word.set('toremove', 'true')
								else:
									word.set('form', merged_form)
									process(lookup(merged_form)['entry'], word)
									next_word.set('toremove', 'true')
									next_next_word.set('toremove', 'true')
							else:
								isunknown = True
					else:
						word.set('form', merged_form)
						process(lookup(merged_form)['entry'], word)
						next_word.set('toremove', 'true')
				else:
					if next_word.get('lacuna') != None:
						merged_form = '%s%s'%(form, next_form)
						if lookup(merged_form)['is_unknown'] == True:
							try:
								next_next_word = words[idx+2]
							except:
								isunknown = True
							else:
								next_next_form = next_next_word.get("form")
								merged_form='%s%s'%(merged_form, next_next_form)
								if lookup(merged_form)['is_unknown'] == True:
									try:
										next_next_next_word = words[idx+3]
									except:
										isunknown = True
									else:
										next_next_next_form = next_next_next_word.get("form")
										merged_form = '%s%s'%(merged_form, next_next_next_form)
										if lookup(merged_form)['is_unknown'] == True:
											isunknown = True
										else:
											word.set('form', merged_form)
											process(lookup(merged_form)['entry'], word)
											next_word.set('toremove', 'true')
											next_next_word.set('toremove', 'true')
											next_next_next_word.set('toremove', 'true')
								else:
									word.set('form', merged_form)
									word.set('lacuna', 'true')
									process(lookup(merged_form)['entry'], word)
									next_word.set('toremove', 'true')
									next_next_word.set('toremove', 'true')
						else:
							word.set('form', merged_form)
							word.set('lacuna', 'true')
							process(lookup(merged_form)['entry'], word)
							next_word.set('toremove', 'true')
					else:
						if form[-1] == '-':
							merged_form = '%s%s'%(form, next_form)
							if lookup(merged_form)['is_unknown'] == True:
								isunknown = True
							else:
								word.set('form', merged_form)
								process(lookup(merged_form)['entry'], word)
								next_word.set('toremove', 'true')
						else:
							isunknown = True
		#####end test lacunae (1)#####
		#####test lacunae (merge with previous word and next word)#####
			if isunknown == True and (word.get('lacuna') != None or previous_word.get('lacuna') != None):
				merged_form='%s%s'%(previous_form, form)
				if lookup(merged_form)['is_unknown'] == True:
					merged_form='%s%s'%(merged_form, next_form)
					if lookup(merged_form)['is_unknown'] == True:
						isunknown = True
					else:
						previous_word.set('form', merged_form)
						previous_word.set('lacuna', 'true')
						process(lookup(merged_form)['entry'], previous_word)
						word.set('toremove', 'true')
						next_word.set('toremove', 'true')
				else:
					previous_word.set('form', merged_form)
					previous_word.set('lacuna', 'true')
					process(lookup(merged_form)['entry'], previous_word)
					word.set('toremove', 'true')
		#####end test lacunae (2)#####

			if isunknown == True:
				document.SubElement(word, "lemma").set("id", "unknown")
		else:
			process(lookup(form)['entry'], word)

	#remove joined words
	toremove=parse.xpath('//word[@toremove]')
	for nTr in toremove:
		nTr.getparent().remove(nTr)
	
	#recalculate indexes
	for sentence in parse.xpath('//sentence'):
		words = sentence.xpath('./word')
		for idx, word in enumerate(words):
			word.set("id", str(idx+1))
			if word.xpath('./lemma')[0].get('id')=='unknown':	
				if word.get('form').find('*') > -1:
					unknownlist_proper.write('%s : %s\n'%(word.get('form'), fileName))
					unknowncount_proper+=1
				elif len(word.get('form')) == 1:
					unknownlist_single.write('%s : %s\n'%(word.get('form'), fileName))
					unknowncount_single+=1
				else:
					unknownlist.write('%s : %s\n'%(word.get('form'), fileName))
					unknowncount+=1
					unknown_dict[cleanWords(word.get('form'))] = ''
	#update spreadsheet		
	record[h['Word count']].value = len(parse.xpath('//word'))
	record[h['Unknown words']].value = unknowncount
	record[h['Unknown proper names']].value = unknowncount_proper
	record[h['Unknown single letters']].value = unknowncount_single
	record[h['Words with multiple lemmata']].value = wml
	record[h['Disambiguable words']].value = wml_mPos
	#save
	parse.write('%s/%s'%(af, fileName), xml_declaration = True, encoding='UTF-8', pretty_print=True)
	wb.save('%s/file_list.xlsx'%file_list)
	print()
unknownlist_single.close()
unknownlist_proper.close()
unknownlist.close()

unknownstring = ''			
for item in unknown_dict.items():
	unknownstring += '%s\n'%item[0]
open('%s/unknown_types.txt'%logf,'w').write(unknownstring)